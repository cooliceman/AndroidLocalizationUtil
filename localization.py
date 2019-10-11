#!/usr/bin/python
# -*- coding: utf-8 -*-
import os, sys, getopt, shutil
import openpyxl
import json
from lxml import etree


def load_xml(path):
    # 创建清除空字符的Parser,否则pretty_print不起作用.
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(path, parser)
    return tree


def exist(tree, key):
    return tree.find("string[@name='%s']" % key) is not None


def main(argv):
    force_update = False
    map_file = None
    reformat = False
    try:
        opts, args = getopt.getopt(argv, "hvrfm:", ["version", "force", "map"])
        for opt, arg in opts:
            if opt == '-h':
                usage()
                sys.exit()
            elif opt == '-r':
                reformat = True
            elif opt in ("-v", "--version"):
                version()
                sys.exit()
            elif opt in ("-f", "--force"):
                force_update = True
            elif opt in ("-m", "--map"):
                map_file = arg
    except getopt.GetoptError:
        usage()
        sys.exit(2)

    try:
        input_file = args[0]
        dest_folder = args[1]
    except:
        usage()
        sys.exit(2)

    if not input_file or not dest_folder:
        usage()
        sys.exit(2)

    if os.path.isfile(input_file) and input_file.endswith("xlsx"):
        print 'Input file:%s\nDestination directory:%s' % (input_file, dest_folder)
        localize(input_file, dest_folder, force_update)
    elif os.path.isdir(input_file):
        if reformat:
            format_file(input_file)

        if map_file is None:
            print 'Source directory:%s\nDestination directory:%s' % (input_file, dest_folder)
            combine_xml(input_file, dest_folder, force_update)
        else:
            print 'Source directory:%s\nDestination directory:%s\nMapping file:%s' % (input_file, dest_folder, map_file)
            combine(input_file, dest_folder, map_file, force_update)
    else:
        usage()
        sys.exit(2)


def usage():
    print """localization.py  [-h] [-v] excel_file_path|translation_folder_path  project_resource_folder  [-f]
    -h  print this usage description
    -v  printh the version
    -f  force override
    """


def version():
    print "version:1.0.0"


def correct_language_code(code):
    """
    校正不规范语言码格式
    :param code: 从文件名中获取到的语言码
    :return: 符合Android规范的语言码
    """

    parts = code.split("_")
    if len(parts) < 2:
        return code

    parts[1] = parts[1].upper()
    return parts[0] + "-r" + parts[1]


def format_file(input_file):
    """
    将文件整理成符合Android工程语言目录结构样式

    :param input_file:待处理数据所在目录
    :return:
    """
    for root, folder_names, file_names in os.walk(input_file):
        for file_name in file_names:
            if file_name.endswith("_strings.xml"):
                # 获取语言前缀
                prefix = file_name.split("_s")[0]
                prefix = correct_language_code(prefix)
                print root, file_name, prefix
                new_dir = root + "/values-" + prefix
                # 创建values目录
                if not os.path.exists(new_dir):
                    os.mkdir(new_dir)
                new_file_path = new_dir + "/" + file_name
                # 移动字符文件
                shutil.move(root + "/" + file_name, new_file_path)
                # 切换目录
                os.chdir(new_dir)
                # 重命名文件
                os.rename(file_name, "strings.xml")
                # 切回上一级目录
                os.chdir("../")


def combine(src_dir, dst_dir, key_map_file, force_update):
    file_list = find_string_resource_files(src_dir)
    dst_file_list = find_string_resource_files(dst_dir)
    with open(os.path.realpath(key_map_file)) as key_map_fp:
        key_map = json.load(key_map_fp)
        # find the target string resource file
        for dst_file in dst_file_list:
            lang_code = os.path.basename(os.path.dirname(dst_file)).replace("values-", "")
            src_file = find_string_resource_file_of(file_list, lang_code)
            if src_file is None:
                print "source file not exist:%s" % lang_code
                continue

            print src_file
            strings = load_xml(src_file)
            dst_strings = load_xml(dst_file)
            rewrite = False
            for key in list(key_map):
                if not exist(strings, key):
                    print "Warning:no key found!!!!"
                    continue

                value = strings.find("string[@name='%s']" % key).text
                if not exist(dst_strings, key_map[key]):
                    # 新建字符结点
                    localize_element = etree.Element('string', name=key_map[key])
                    localize_element.text = value
                    # 追加结点
                    dst_strings.getroot().append(localize_element)
                    rewrite = True
                elif force_update:
                    print "Key already exist, update"
                    dst_strings.find("string[@name='%s']" % key_map[key]).text = value
                    rewrite = True
                else:
                    print "I:%s already exist, skip" % key_map[key]

            if rewrite:
                dst_strings.write(dst_file, xml_declaration=True, encoding='UTF-8', pretty_print=True)


def combine_xml(src_folder, dest_folder, force_update):
    # find string resource file in source folder
    src_file_list = find_string_resource_files(src_folder)
    if not src_file_list:
        print "No resource file found in source folder"
        return
    # find string resource file in destination folder
    dest_file_list = find_string_resource_files(dest_folder)
    if not dest_file_list:
        print "No resource file found in destination folder"
        return
    for src_file in src_file_list:
        values_folder = os.path.split(os.path.dirname(src_file))[1]
        target_file = find_string_resource_file_of(dest_file_list, values_folder)
        if not target_file:
            continue
        print "start combine:%s and %s" %(src_file, target_file)
        src_document = load_xml(src_file)
        dest_document = load_xml(target_file)
        rewrite = False
        # iterator the source xml document element and update the destination xml document
        for entity in src_document.findall("string"):
            key = entity.get("name")
            value = entity.text
            if not key:
                print "Warning:no key found!!!!"
                continue
            else:
                key = key.strip()
            if not value:
                print "I:[%s]no value found, skip" % key
                continue
            else:
                value = value.strip()
            if exist(dest_document, key):
                if force_update:
                    print "Key already exist, update"
                    dest_document.find("string[@name='%s']" % key).text = value
                    rewrite = True
                    continue
                else:
                    print "I:%s already exist, skip" % key
                    continue
            # 新建字符结点
            localize_element = etree.Element('string', name=key)
            localize_element.text = value
            # 追加结点
            dest_document.getroot().append(localize_element)
            rewrite = True
        if rewrite:
            dest_document.write(target_file, xml_declaration=True, encoding='UTF-8', pretty_print=True)


def localize(input_file, dest_dir, force_update):
    # 打开多语言翻译excel文件
    try:
        wb = openpyxl.load_workbook(input_file)
    except:
        print "Excel file can't be open"
        return
    # 获取默认Sheet
    ws = wb.active
    # 获取字符串资源文件列表
    file_list = find_string_resource_files(dest_dir)
    if not file_list:
        print "No resource file found in destination folder"
        return
    # 遍历Excel每个语言数据行
    for col in range(3, ws.max_column + 1):
        # 提取语言码
        lang_code = ws.cell(row=1, column=col).value
        if not lang_code:
            print "no language code, skip"
            continue
        else:
            lang_code = lang_code.strip()
        # 获取对应语言的资源文件
        lang_string_file = find_string_resource_file_of(file_list, "values-" + lang_code)
        if not lang_string_file:
            continue
        string_map = load_xml(lang_string_file)
        print "\n%s %s" % (lang_code, lang_string_file)
        rewrite = False #标识文件是否有变更
        for r in range(4, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            value = ws.cell(row=r, column=col).value
            print key, ":", value
            if not key:
                print "no key exist at:%d,%d" % (r, col)
                break
            else:
                key = key.strip()
            if not value:
                print "no value found, skip"
                continue
            else:
                value = value.strip()
            if exist(string_map, key):
                if force_update:
                    print "Key already exist, update"
                    string_map.find("string[@name='%s']" % key).text = value
                    rewrite = True
                    continue
                else:
                    print "Warning:%s already exist, skip" % key
                    continue
            # 新建字符结点
            localize_element = etree.Element('string', name=key)
            localize_element.text = value
            # 追加结点
            string_map.getroot().append(localize_element)
            rewrite = True
        if rewrite:
            string_map.write(lang_string_file, xml_declaration=True, encoding='UTF-8', pretty_print=True)


def find_string_resource_files(folder):
    """

    遍历目录获取所有语言的string.xml文件
    """
    file_list = []
    for root, dirs, files in os.walk(os.path.abspath(folder)):
        for d in dirs:

            find_string_resource_files(d)
            if "res" == d:
                break

        for f in files:
            if f == "strings.xml":
                file_list.append(os.path.join(root, f))
    return file_list


def find_string_resource_file_of(file_list, values_dir_name):
    """find_string_resource_file_of(file_list, values_dir_name)

    通过语言码对应文件夹名称来找到对应的string.xml文件路径
    """
    for f in file_list:
        if os.path.dirname(f).endswith(values_dir_name):
            return f


if __name__ == "__main__":
    main(sys.argv[1:])
    exit(0)
